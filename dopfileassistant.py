import os
import logging
from datetime import datetime

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


# Configure the logging settings
logging.basicConfig(filename='doplogs.log' ,level=20, format='%(asctime)s - %(levelname)s - %(message)s')


class FileAssistantError(Exception):
    pass

# Does not take any argument while initialize
class DOPFileAssistant:
    # Takes path of a xlsx file and format it and save to same path
    def format_excel_file(self,input_path):
        try:
            # Load workbook
            wb = openpyxl.load_workbook(input_path)
            sheet = wb.active
            logging.info("Load Workbook Sucessful !")

            # Change size of all text to 9 and align vertically to center
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(vertical='center')
                    
            logging.info("Font Size Change Sucessful !")

            # Make height of all rows to approximately 25 pixels (15 points)
            for row in sheet.iter_rows():
                sheet.row_dimensions[row[0].row].height = 15

            # Make text in 1st, 6th and last two rows bold and change their background color to grey
            for row in [1, 6, sheet.max_row, sheet.max_row - 1]:
                for cell in sheet[row]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCFC", end_color="CCCCFC", fill_type="solid")

            # Change column width
            column_widths_pixels = [35, 90, 100, 190, 100, 100, 60, 60, 60]
            column_widths_points = [i / 9.5 for i in column_widths_pixels]  # Converted from pixels to points
            for i, column_width_points in enumerate(column_widths_points, start=1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width_points

            # Apply wrap text to row 6
            for cell in sheet[6]:
                cell.alignment = Alignment(wrap_text=True, vertical='center')

            # Change height of row 6 to approximately 90 pixels (54 points)
            sheet.row_dimensions[6].height = 54
            logging.info("Coloring and Formatting Sucessful !")

            # Save workbook
            wb.save(input_path)
            logging.info("Report file edited and Saved Sucessful !")
        except Exception as e :
            logging.error(f"Error Occoured While Formatting Excel File : {e}")
            raise FileAssistantError(f"Error Occoured While Formatting Excel File : {e}")

    # Takes path of a xls file and save xlsx file at same path
    def extract_xlsx_file(self,file_path):
        try:
            # Read the Excel file
            df = pd.read_excel(file_path, engine='xlrd')
            logging.info("xls file read Sucessful !")

            # Set columns to range of shape
            df.columns = range(df.shape[1])

            # Drop specified columns
            df = df.drop(columns=[0, 4, 11, 15, 16, 17, 18, 19, 20, 21, 22], axis=1)

            # Delete rows 6 to 10
            df.drop(df.index[6:11], inplace=True)

            # Move contents of column 8 and row 2,3,4 to column 2 and delete column 8
            df.loc[2:5, 3] = df.loc[2:5, 8]
            df.drop(8, axis=1, inplace=True)

            # Copy values from column 1 to column 2 for the last two rows
            df.iloc[-2:, 2] = df.iloc[-2:, 1].values

            # Drop column with label 2
            df.drop(2, axis=1, inplace=True)

            # Copy values from column 1 to column 2 for the last two rows
            df.iloc[-2:, 6] = df.iloc[-2:, 5].values

            # Drop column with label 2
            df.drop(9, axis=1, inplace=True)

            # Insert text 'No.' in the cell at the 11th index of column 1
            df.loc[11, 1] = 'No'

            # Update values in the first column
            df.iloc[7:-2, 0] = range(1, len(df.iloc[7:-2, 0]) + 1)

            # Drop row with index 1
            df.drop(1, inplace=True)

            # Update values in columns 1:3 for the first five rows
            df.iloc[:5, 1:3] = df.iloc[:5, :2]
            df.iloc[:5, 0] = np.nan

            # Update values in columns 2:3 for the first five rows
            df.iloc[:5, 3] = df.iloc[:5, 2]
            df.iloc[:5, 2] = np.nan
            logging.info("Edited File Sucessful !")
            
            # Check if the file exists
            if os.path.isfile(file_path):
                # Use os.remove() to delete the file
                os.remove(file_path)
                
            # Save the modified DataFrame to a new Excel file
            df.to_excel(file_path.replace('.xls','.xlsx'), header=False, index=False)
            logging.info("File Saved as xlsx Sucessful !")
        except Exception as e:
            logging.error(f"Error Occoured while saving file as xlsx : {e}")
            raise FileAssistantError(f"Error Occoured while saving file as xlsx : {e}")

    # Takes a path where we want to download files and create folder structure there and returns records and declaration path
    def create_directories_and_get_paths(self, download_folder):
        try:
            # Get the current date
            current_date = datetime.now()
            logging.info("Got Current date and time Sucessful !")

            # Format the date in "MM-YYYY" and "DD-MM-YYYY" format
            month_year = current_date.strftime("%m-%Y")
            day_month_year = current_date.strftime("%d-%m-%Y")

            # Get the path of the current python file
            current_file_path = download_folder

            # Create the "RDReports" directory if it doesn't exist
            rdreports_path = os.path.join(current_file_path, "RDReports")
            os.makedirs(rdreports_path, exist_ok=True)
            logging.info("RDReports Directory created Sucessful !")

            # Create the "MM-YYYY" directory inside "RDReports" if it doesn't exist
            month_year_path = os.path.join(rdreports_path, month_year)
            os.makedirs(month_year_path, exist_ok=True)
            logging.info("Month directory created Sucessful !")

            # Create the "DD-MM-YYYY" directory inside "MM-YYYY" if it doesn't exist
            day_month_year_path = os.path.join(month_year_path, day_month_year)
            os.makedirs(day_month_year_path, exist_ok=True)
            logging.info("Date directory created Sucessful !")

            # Create the "Reports" and "Declarations" directories inside "DD-MM-YYYY" if they don't exist
            report_path = os.path.join(day_month_year_path, "Reports")
            os.makedirs(report_path, exist_ok=True)

            declaration_path = os.path.join(day_month_year_path, "Declarations")
            os.makedirs(declaration_path, exist_ok=True)
            logging.info("Reports and Declarations Directory created Sucessful !")

            return report_path, declaration_path
        except Exception as e:
            logging.error(f"Error while creating directory and get path : {e}")
            raise FileAssistantError(f"Error while creating directory and get path : {e}")
    
