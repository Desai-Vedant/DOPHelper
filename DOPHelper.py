import sys
import os
import re
import json
import tempfile
import webbrowser
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from PySide6.QtWidgets import (
    QApplication, QFileDialog, QWidget, QVBoxLayout, QDialog, QHBoxLayout,
    QComboBox, QLabel, QLineEdit, QLayout, QPushButton, QTableWidgetItem, 
    QAbstractItemView, QHeaderView, QTableWidget, QMessageBox, QFrame, QMainWindow,
    QStackedWidget, QCheckBox, QTextEdit
)
from PySide6.QtCore import (
    Qt, Signal, QThread
)
from PySide6.QtGui import QIcon

from dopdatabaseassistant import DOPDatabaseAssistant
from dopwebassistant import DOPWebAssistant
from dopfileassistant import DOPFileAssistant

from qt_material import apply_stylesheet  # Ensure this is imported after PySide6




# BASE UI CLASS
class RDApplication(QMainWindow):

    def __init__(self):
        super().__init__()
        self.setWindowIcon(QIcon('./_internal/static/icon.svg'))
        # Get the screen geometry
        screen = QApplication.primaryScreen()
        screen_geometry = screen.geometry()
        
        # Set the window size to the screen's width and height
        self.resize(screen_geometry.width(), screen_geometry.height())

        self.user_id = "USERID"
        self.user_password = "PASSWORD"
        self.def_download_dir = "./"
        self.agent_id = "AGENTID"
        self.lisence_exp_date = "LISENCEEXP"
        self.agent_address = 'AGENTADDRESS'
        self.agent_name = 'AGENTNAME'
        self.agent_husband_name = 'AGENTHUSBANDNAME'
        self.ocr_apikey = "APIKEY"
        self.theme = "Dark"
        self.ascent = "amber"
        self.scale = "0"
        self.initialize_settings_json()
        self.load_settings()
        self.initialize_assistants()
        self.init_ui()

    # Method to Initialize modules
    def initialize_assistants(self):
        self.dda = DOPDatabaseAssistant()
        self.dwa = DOPWebAssistant()
        self.dfa = DOPFileAssistant()
        self.dwa.user_id = self.user_id
        self.dwa.user_password = self.user_password

    # Define UI
    def init_ui(self):
        # Setting up Window Properties
        self.setWindowTitle("DOP Helper Application")

        # Setting up Central Widget
        self.central_widget = QWidget()
        self.central_widget.setObjectName("centralWidget")  # Set the object name for CSS targeting
        self.central_widget.setContentsMargins(0,0,0,0)
        self.setCentralWidget(self.central_widget)

        # Create a frame for the sidebar
        self.sidebar_frame = QFrame()
        self.sidebar_frame.setObjectName("sidebarFrame")  # Set the object name for CSS targeting

        # Button Layout for Holding Buttons with 10px side spacing
        self.sidebar_buttons_layout = QVBoxLayout()
        self.sidebar_buttons_layout.setContentsMargins(10,10,10,10)

        # QStack Widget to hold Pages
        self.stacked_widget = QStackedWidget()
        self.stacked_widget.setObjectName("stackedWidget")  # Set the object name for CSS targeting
        self.stacked_widget.setContentsMargins(0,0,0,0)


        # Layout for Sidebar which is placed in sidebar frame and holds buttons layout
        self.sidebar_layout = QVBoxLayout(self.sidebar_frame)
        self.sidebar_layout.setAlignment(Qt.AlignTop)  # Align the buttons to the top
        self.sidebar_layout.addLayout(self.sidebar_buttons_layout)
        self.sidebar_layout.setContentsMargins(0,0,0,0)

        # Creating Layout for central widget
        self.main_layout = QHBoxLayout()
        self.main_layout.setSpacing(0)
        self.main_layout.addWidget(self.sidebar_frame, 1)
        self.main_layout.addWidget(self.stacked_widget, 4)
        self.main_layout.setContentsMargins(5,5,5,5)

        # Load and create all pages
        self.pages = [
            DashboardPage(self),
            WorkspacePage(self),
            ViewAccountsPage(self),
            AslaasUpdatePage(self),
            DownloadReportPage(self),
            SettingsPage(self),
            AboutPage(self)
        ]

        # Setting up Current Page to DashboardPage
        self.current_page = 0
        self.stacked_widget.setCurrentIndex(self.current_page)

        # Setting up layout for central widget
        self.central_widget.setLayout(self.main_layout)

        # Create and add the appearance button to open AppearanceWindow
        appearance_button = QPushButton("Appearance")
        appearance_button.clicked.connect(self.open_appearance_window)
        appearance_button.setIcon(QIcon("./_internal/static/appearance.svg"))
        self.sidebar_buttons_layout.addWidget(appearance_button)
        self.appearance_window = AppearanceWindow(self)

    # Open Appearance Window
    def open_appearance_window(self):
        self.appearance_window.show()

    # Close Appearance window on closing main window
    def closeEvent(self, event):
        if self.appearance_window is not None:
            self.appearance_window.close()
        event.accept()

    # Reguler popup message for info
    def show_popup_message(self, title, message, icon=QMessageBox.Information):
        msg_box = QMessageBox(self)
        msg_box.setIcon(icon)
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        msg_box.exec()

    # Popup Message for error
    def show_error_message(self, title, message):
        self.show_popup_message(title, message, QMessageBox.Critical)

   # Method to create settings.json if not already exists
    def initialize_settings_json(self):
        settings_data = {
            'user_id': "USERID",
            'user_password': "PASSWORD",
            'def_download_dir': "./",
            'agent_id': "AGENTID",
            'lisence_exp_date': "EXPDATE",
            'agent_address': "ADDRESS",
            'agent_name': "AGENTNAME",
            'agent_husband_name': "AGENTHUSBANDNAME",
            'ocr_apikey':"APIKEY",
            "theme": "Dark",
            "ascent": "blue",
            "scale": "0"
        }

        try:
            with open('settings.json', 'x') as file:
                json.dump(settings_data, file)
        except FileExistsError:
            pass

    # Method to save the settings to file
    def save_settings(self, filename='settings.json'):
        settings_data = {
            'user_id': self.user_id,
            'user_password': self.user_password,
            'def_download_dir': self.def_download_dir,
            'agent_id': self.agent_id,
            'lisence_exp_date': self.lisence_exp_date,
            'agent_address': self.agent_address,
            'agent_name': self.agent_name,
            'agent_husband_name': self.agent_husband_name,
            'ocr_apikey' : self.ocr_apikey,
            "theme": self.theme,
            "ascent": self.ascent,
            "scale": self.scale
        }

        with open(filename, 'w') as file:
            json.dump(settings_data, file)

    # Method to load the settings from file
    def load_settings(self, filename='settings.json'):
        try:
            with open(filename, 'r') as file:
                settings_data = json.load(file)

            self.user_id = settings_data.get('user_id', '')
            self.user_password = settings_data.get('user_password', '')
            self.def_download_dir = settings_data.get('def_download_dir', '')
            self.agent_id = settings_data.get('agent_id', '')
            self.lisence_exp_date = settings_data.get('lisence_exp_date','')
            self.agent_address = settings_data.get('agent_address','')
            self.agent_name = settings_data.get('agent_name', '')
            self.agent_husband_name = settings_data.get('agent_husband_name', '')
            self.ocr_apikey = settings_data.get('ocr_apikey','')
            self.theme = settings_data.get('theme','')
            self.ascent = settings_data.get('ascent','')
            self.scale = settings_data.get('scale','')
        except FileNotFoundError:
            # Handle the case where the file is not found (e.g., first-time setup)
            pass

    # Method to create sidebar button with icon
    def create_sidebar_button(self, text, index, icon_path=""):
        button = QPushButton(text)
        button.setObjectName("sidebarButton")  # Set object name for CSS targeting
        button.setIcon(QIcon(icon_path))  # Set the icon for the button
        button.clicked.connect(lambda: self.change_page(index))
        self.sidebar_buttons_layout.addSpacing(0)  # Add custom spacing between buttons
        self.sidebar_buttons_layout.addWidget(button)

    # Method to change the page
    def change_page(self, index):
        self.stacked_widget.setCurrentIndex(index)


# DOWNLOAD REPORT THREAD
class DownloadReportThread(QThread):
    finished_signal = Signal(str)
    finished_with_error = Signal(str)

    def __init__(self, parent, lot_reference):
        super().__init__()
        self.parent = parent
        self.lot_reference = lot_reference

    def run(self):
        try:
            reports_path, dec_path = self.parent.dfa.create_directories_and_get_paths(self.parent.def_download_dir)
            self.parent.dwa.open_browser_portal()
            self.parent.dwa.login()
            new_path = self.parent.dwa.perform_download_report_task(self.lot_reference, reports_path)
            self.parent.dfa.extract_xlsx_file(new_path)
            self.parent.dfa.format_excel_file(new_path.replace(".xls", ".xlsx"))
            self.finished_signal.emit(new_path.replace(".xls", ".xlsx"))
        except Exception as e:
            self.finished_with_error.emit(str(e))
        finally:
            self.parent.dwa.close_browser()


# ASLAAS UPDATE THREAD
class AslaasUpdateThread(QThread):
    finished_signal = Signal()
    finished_with_error = Signal(str)

    def __init__(self, parent, acc_nos, aslaas_nos, acc_ids):
        super().__init__()
        self.parent = parent
        self.acc_nos = acc_nos
        self.aslaas_nos = aslaas_nos
        self.acc_ids = acc_ids

    def run(self):
        try:
            if self.acc_nos:
                self.parent.dwa.open_browser_portal()
                self.parent.dwa.login()
                self.parent.dwa.perform_update_aslaas_task(self.acc_nos, self.aslaas_nos)
                self.parent.dda.sync_aslaas_numbers(self.acc_ids, self.aslaas_nos)
            self.finished_signal.emit()
        except Exception as e:
            self.finished_with_error.emit(str(e))
        finally:
            self.parent.dwa.close_browser()


# SYNC ACCOUNTS THREAD
class SyncAccountsThread(QThread):
    finished_signal = Signal()
    finished_with_error = Signal(str)

    def __init__(self, parent):
        super().__init__()
        self.parent = parent

    def run(self):
        try:
            self.parent.dwa.open_browser_portal()
            self.parent.dwa.login()
            self.parent.dwa.download_accounts_list_task()
            self.parent.dwa.download_aslaas_csv()
            self.parent.dda.sync_database_task()
            self.finished_signal.emit()
        except Exception as e:
            self.finished_with_error.emit(str(e))
        finally:
            self.parent.dwa.close_browser()


# PERFORM LOT THREAD
class PerformLotThread(QThread):
    finished_signal = Signal(str)
    finished_with_error = Signal(str)

    def __init__(self, parent, acc_nos, acc_ins, reports_path):
        super().__init__()
        self.parent = parent
        self.acc_nos = acc_nos
        self.acc_ins = acc_ins
        self.reports_path = reports_path

    def run(self):
        try:
            if self.acc_nos:
                self.parent.dwa.open_browser_portal()
                self.parent.dwa.login()
                ref_no = self.parent.dwa.perform_lot_task(self.acc_nos,self.acc_ins)
                new_path = self.parent.dwa.perform_download_report_task(ref_no, self.reports_path)
                self.parent.dfa.extract_xlsx_file(new_path)
                self.parent.dfa.format_excel_file(new_path.replace(".xls",".xlsx"))
            self.finished_signal.emit(self.reports_path)
        except Exception as e:
            self.finished_with_error.emit(str(e))
        finally:
            self.parent.dwa.close_browser()


# Dashboard for Agent
class DashboardPage(QWidget):
    def __init__(self, parent):
        super().__init__()
        self.parent = parent
        self.init_ui()

    def init_ui(self):
        # Add widgets and layout for DashboardPage
        layout = QVBoxLayout()
        layout.addWidget(QLabel("WELCOME"), alignment=Qt.AlignCenter)
        self.setLayout(layout)
        self.parent.stacked_widget.addWidget(self)

        self.parent.create_sidebar_button("Dashboard", 0, './_internal/static/dashboard.svg')


# Get Account Numbers and Declaration
class WorkspacePage(QWidget):
    def __init__(self, parent):
        super().__init__()
        self.parent = parent
        self.account_ids = []
        self.account_nos = []
        self.account_inst = []
        self.init_ui()

    def init_ui(self):
        # Widgets for WorkspacePage
        self.client_id_label = QLabel("Enter Client IDs (separated by comma):")
        self.client_id_edit = QLineEdit()

        self.get_account_numbers_button = QPushButton("Get Account Numbers")
        self.get_account_numbers_button.clicked.connect(self.get_account_numbers)

        self.account_numbers_label = QLabel("Account Numbers:")
        self.account_numbers_text_edit = QTextEdit()
        self.account_numbers_text_edit.setReadOnly(True)

        self.view_details_button = QPushButton("View Details")
        self.view_details_button.clicked.connect(self.view_details)

        self.perform_lot_button = QPushButton("Perform Lot and Download Report")
        self.perform_lot_button.clicked.connect(self.perform_lot)

        self.download_declaration_button = QPushButton("Download Declaration")
        self.download_declaration_button.clicked.connect(self.download_declaration)

        # Layout for WorkspacePage
        layout = QVBoxLayout()
        layout.addWidget(self.client_id_label)
        layout.addWidget(self.client_id_edit)
        layout.addWidget(self.get_account_numbers_button)
        layout.addWidget(self.account_numbers_label)
        layout.addWidget(self.account_numbers_text_edit)
        layout.addWidget(self.view_details_button)
        layout.addWidget(self.perform_lot_button)
        layout.addWidget(self.download_declaration_button)

        self.setLayout(layout)
        self.parent.stacked_widget.addWidget(self)

        self.parent.create_sidebar_button("Workspace", 1, "./_internal/static/workspace.svg")

    def get_account_numbers(self):
        try:
            # Get the entered client IDs separated by commas
            client_ids_str = self.client_id_edit.text()

            # Get active account id's
            active_acc_ids = self.parent.dda.get_list_of_active_account_ids()

            # Split the input into a list of individual client IDs
            client_code = client_ids_str.split(",")
            acc_ids = []
            acc_inst = []
            for code in client_code:
                ac_data = code.split('-')
                if str(ac_data[0]) in active_acc_ids:
                    acc_ids.append(str(ac_data[0]))
                    if len(ac_data)==2 and not ac_data[1]=="":
                        acc_inst.append(ac_data[1])
                    else:
                        acc_inst.append(1)

            if not len(acc_ids)==0:
                # Combine the two lists into a list of tuples
                combined = list(zip(acc_ids, acc_inst))

                # Sort the combined list
                combined.sort()

                # Now, 'combined' is sorted based on the first element of each tuple
                # You can get the sorted 'acc_ids' and its corresponding 'acc_inst' like this:
                acc_ids, acc_inst = zip(*combined)

            self.account_ids = acc_ids
            self.account_inst = acc_inst

            # Retrieve the corresponding account numbers from the dictionary
            account_numbers = self.parent.dda.get_acc_nos_using_ids(acc_ids)

            self.account_nos = account_numbers

            # Display the account numbers in the textbox
            self.account_numbers_text_edit.setPlainText(", ".join(account_numbers))
        except:
            self.parent.show_error_message("Error", "Enter Valid Account ID's")

    def view_details(self):
        try:
            acc_nos = self.account_nos
            acc_ids = self.account_ids
            acc_ins = self.account_inst

            acc_ids = list(map(int, acc_ids))
            acc_ins = list(map(int, acc_ins))

            df = self.parent.dda.get_ac_details_by_ids(acc_ids)

            # Convert the list to a pandas Series
            values_series = pd.Series(acc_ins,name='installments')
            df['installments'] = acc_ins

            # Multiply the 'denomination' column with the values
            df.denomination = df.denomination.astype(int) * df.installments.astype(int)

            # Create a new column with sequential numbering starting from 1
            df.insert(0, 'sr_no', range(1, len(df) + 1))

            total_amount = df.denomination.sum()
            df.loc[len(df.index)] = ["", "", "", 'Total Amount : ', total_amount, "", ""]

            # Display the account numbers and client details in a pop-up window
            self.show_client_details_popup(df)
        except:
            self.parent.show_error_message("Error", "Error while showing Error Message")

    def show_client_details_popup(self, df):
        # Create a pop-up window to display the client details
        popup = QDialog(self)
        popup.setWindowTitle("Client Details")
        popup.setGeometry(200, 200, 650, 350)
        popup.setStyleSheet(self.parent.styleSheet())  # Use the same stylesheet as the main window

        layout = QVBoxLayout()

        # Create a table to display the client details
        table_widget = QTableWidget()
        table_widget.setColumnCount(len(df.columns))
        table_widget.setRowCount(len(df))

        # Set the column names
        column_names = ["No", "Acc. No", "Acc. Id", "Name", "Amount", "Open Date","Installments"]
        table_widget.setHorizontalHeaderLabels(column_names)

        for row in range(len(df)):
            for col in range(len(df.columns)):
                table_widget.setItem(row, col, QTableWidgetItem(str(df.iloc[row, col])))

        # Adjust the column widths and set stretch factors for the "View Clients" page table
        table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        table_widget.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        table_widget.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        table_widget.setColumnWidth(1, 120)  # Set a fixed width for the RD Account Number column

        # Set the edit triggers to disable editing
        table_widget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        table_widget.verticalHeader().setVisible(False)


        layout.addWidget(table_widget)

        popup.setLayout(layout)
        popup.exec()

    def perform_lot(self):
        try:
            acc_nos = self.account_nos
            acc_ids = self.account_ids
            acc_ins = self.account_inst

            acc_ids = list(map(int, acc_ids))
            acc_ins = list(map(int, acc_ins))

            reports_path, dec_path = self.parent.dfa.create_directories_and_get_paths(self.parent.def_download_dir)

            # Create and start the PerformLotThread
            self.perform_lot_thread = PerformLotThread(self.parent, acc_nos, acc_ins, reports_path)
            self.perform_lot_thread.finished_signal.connect(self.perform_lot_completed)
            self.perform_lot_thread.finished_with_error.connect(self.perform_lot_error)
            self.perform_lot_thread.start()
        except:
            self.parent.show_error_message("Error","Error While Performing Lot")

    def perform_lot_completed(self, path):
        # Add any additional UI updates or operations needed after the download completes
        self.parent.show_popup_message("Message", f"Lot Performed Sucessfuly. Downloaded report at: {path}")

    def perform_lot_error(self, error_message):
        self.parent.show_error_message("Error" , "Error While Performing Lot")

    def download_declaration(self):
        try:
            acc_ids = self.account_ids
            acc_ids = list(map(int, acc_ids))
            df = self.parent.dda.get_data_for_declaration(acc_ids=acc_ids)
            acc_details = df.to_numpy()

            wb = load_workbook('./_internal/static/POSTDECLARE.xlsx')
            ws = wb.active

            report_path,declaration_path = self.parent.dfa.create_directories_and_get_paths(self.parent.def_download_dir)

            # Get the current date
            current_date = datetime.now()

            # Format the new file name
            new_file_name = f"Declaration-{current_date.strftime('%d-%m-%Y-%H-%M-%S')}.xlsx"

            # Get the new file path
            new_file_path = os.path.join(declaration_path, new_file_name)

            wb.save(new_file_path)
            wb2 = load_workbook(new_file_path)
            ws2 = wb2.active

            for i in range(0,len(acc_details)):
                #ws2['A'+str(10+i+1)] = acc_details[i][0]
                ws2['B'+str(9+i+1)] = acc_details[i][0]
                ws2['C'+str(9+i+1)] = acc_details[i][1]
                ws2['D'+str(9+i+1)] = acc_details[i][2]
                ws2['E'+str(9+i+1)] = acc_details[i][3]

            for x in range((9+len(acc_details)+1),55):
                ws2.row_dimensions[x].hidden = True

            # Additional functionality to replace specific text values
            agent_id_text = "agent_id"
            agent_name_text = "agent_name"
            agent_husband_name_text = "agent_husband_name"
            agent_address_text = "agent_address"
            license_exp_date_text = "lisence_exp_date"

            texts = [agent_id_text, agent_name_text, agent_husband_name_text, agent_address_text, license_exp_date_text]

            agent_id_value = str(self.parent.agent_id)  # Convert to string to handle None values
            agent_name_value = str(self.parent.agent_name)
            agent_husband_name_value = str(self.parent.agent_husband_name)
            agent_address_value = str(self.parent.agent_address)
            license_exp_date_value = str(self.parent.lisence_exp_date)

            values = [agent_id_value, agent_name_value, agent_husband_name_value, agent_address_value, license_exp_date_value]

            a7text = ws2['A7'].value
            a7text = self.replace_text_in_cell(a7text, agent_name_text, agent_name_value)
            a7text = self.replace_text_in_cell(a7text, agent_id_text, agent_id_value)
            a7text = self.replace_text_in_cell(a7text, agent_husband_name_text, agent_husband_name_value)
            a7text = self.replace_text_in_cell(a7text, agent_address_text, agent_address_value)
            a7text = self.replace_text_in_cell(a7text, license_exp_date_text, license_exp_date_value)
            ws2['A7'] = a7text

            a59text = ws2['A59'].value
            a59text = self.replace_text_in_cell(a59text, agent_id_text, agent_id_value)
            ws2['A59'] = a59text

            wb2.save(new_file_path)
            self.parent.show_popup_message("Message",f"Declaration is Downloaded at : {new_file_path}")
        except Exception as e:
            self.parent.show_error_message("Error","Something went wrong while downloading declaration")

    def replace_text_in_cell(self, cell_value, old_text, new_text):
        # Use regular expression to find and replace the old text
        return re.sub(re.escape(old_text), new_text, str(cell_value))

# View all Accounts
class ViewAccountsPage(QWidget):
    def __init__(self, parent):
        super().__init__()
        self.parent = parent
        self.account_data = self.parent.dda.get_all_accounts()
        self.init_ui()

    def init_ui(self):
        # Create widgets for ViewClientsPage
        self.client_table_widget = QTableWidget()
        self.client_table_widget.setColumnCount(8)
        # Column Names : ["Client ID", "RD Account Number", "Name", "Deposit Amount", "Months Paid","Opening Date","Account Status","Aslaas No."]
        self.client_table_widget.setHorizontalHeaderLabels(["ID", "Acc. No.", "Name", "Deposit", "Months","Opening Date","Status","Aslaas"])

        # Adjust the column widths and set stretch factors for the "View Clients" page table
        self.client_table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.client_table_widget.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.client_table_widget.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.client_table_widget.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.client_table_widget.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.client_table_widget.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.client_table_widget.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        self.client_table_widget.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeToContents)
        #self.client_table_widget.setColumnWidth(1, 120)  # Set a fixed width for the RD Account Number column
        self.client_table_widget.setRowHeight(0, 50)  # Set a fixed height for the Title row

        # Print Button
        self.print_button = QPushButton("Print")
        self.print_button.clicked.connect(self.print_accounts)

        # Refresh Button
        self.refresh_button = QPushButton("Refresh")
        self.refresh_button.clicked.connect(self.refresh_accounts)

        # Layout for ViewClientsPage
        layout = QVBoxLayout()
        layout.addWidget(self.client_table_widget)
        layout.addWidget(self.refresh_button)
        layout.addWidget(self.print_button)

        self.setLayout(layout)
        self.parent.stacked_widget.addWidget(self)
        self.clear_table()
        self.populate_table(self.account_data)

        # Set the edit triggers to disable editing
        self.client_table_widget.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.client_table_widget.verticalHeader().setVisible(False)

        self.parent.create_sidebar_button("View Accounts", 2, './_internal/static/viewaccounts.svg')

    def add_client_to_table(self, client_id, rd_account_number, name, denomination, installments, date, is_active, aslaas_no):
        # Add the client to the table on the View Clients page
        row_count = self.client_table_widget.rowCount()
        self.client_table_widget.insertRow(row_count)
        self.client_table_widget.setItem(row_count, 0, QTableWidgetItem(str(client_id)))
        self.client_table_widget.setItem(row_count, 1, QTableWidgetItem(rd_account_number))
        self.client_table_widget.setItem(row_count, 2, QTableWidgetItem(name))
        self.client_table_widget.setItem(row_count, 3, QTableWidgetItem(denomination))
        self.client_table_widget.setItem(row_count, 4, QTableWidgetItem(installments))
        self.client_table_widget.setItem(row_count, 5, QTableWidgetItem(date))
        self.client_table_widget.setItem(row_count, 6, QTableWidgetItem(is_active))
        self.client_table_widget.setItem(row_count, 7, QTableWidgetItem(aslaas_no))

    def clear_table(self):
        # Clear all rows from the table on the View Clients page
        self.client_table_widget.setRowCount(0)

    def populate_table(self, account_data):
        # Clear existing rows
        self.client_table_widget.setRowCount(0)

        # Add new rows with account data
        for index, row in account_data.iterrows():
            self.add_client_to_table(str(row['ac_id']), str(row['ac_no']), str(row['acc_holder_name']), str(row['denomination']), str(row['no_of_installments']), str(row['acc_opening_date']), str(row['is_active']), str(row['aslaas_no']))

    def refresh_accounts(self):
        self.account_data = self.parent.dda.get_all_accounts()
        self.clear_table()
        self.populate_table(self.account_data)

    def print_accounts(self):
        df = self.parent.dda.get_all_accounts()
        # Save DataFrame to a temporary HTML file
        with tempfile.NamedTemporaryFile(suffix=".html", dir="temp", delete=False) as temp_html_file:
            df.to_html(temp_html_file.name)

            # Open the HTML file in the default web browser
            webbrowser.open('file://' + os.path.realpath(temp_html_file.name))


# Deposit Amount
class AslaasUpdatePage(QWidget):
    def __init__(self, parent):
        super().__init__()
        self.parent = parent
        self.account_ids = []
        self.account_nos = []
        self.aslaas_numbers = []
        self.init_ui()

    def init_ui(self):
        # widgets and layout for AslaasUpdatePage
        self.client_id_label = QLabel("Enter Client IDs to Update Aslaas (separated by comma):")
        self.client_id_edit = QLineEdit()

        self.get_update_details_button = QPushButton("Get Update Details")
        self.get_update_details_button.clicked.connect(self.get_update_details)

        self.account_numbers_label = QLabel("Account Numbers:")
        self.account_numbers_text_edit = QTextEdit()
        self.account_numbers_text_edit.setReadOnly(True)

        self.confirm_update_button = QPushButton("Confirm Updates")
        self.confirm_update_button.clicked.connect(self.confirm_updates)

        layout = QVBoxLayout()
        layout.addWidget(self.client_id_label)
        layout.addWidget(self.client_id_edit)
        layout.addWidget(self.get_update_details_button)
        layout.addWidget(self.account_numbers_label)
        layout.addWidget(self.account_numbers_text_edit)
        layout.addWidget(self.confirm_update_button)

        self.setLayout(layout)
        self.parent.stacked_widget.addWidget(self)

        self.parent.create_sidebar_button("Update Aslaas Numbers", 3, "./_internal/static/update.svg")

    def get_update_details(self):
        try:
            # Get the entered client IDs separated by commas
            client_ids_str = self.client_id_edit.text()
            self.account_numbers_text_edit.clear()

            aslaas_nos = []
            acc_ids = []

            # Get active account id's
            active_acc_ids = self.parent.dda.get_list_of_active_account_ids()

            # Split the input into a list of individual client IDs
            client_code = client_ids_str.split(",")
            for code in client_code:
                ac_data = code.split('-')
                if str(ac_data[0]) in active_acc_ids:
                    acc_ids.append(str(ac_data[0]))
                    if len(ac_data)==2 and not ac_data[1]=="":
                        aslaas_nos.append(ac_data[1])
                    else:
                        aslaas_nos.append("APPLIEDFOR")

            if not len(acc_ids)==0:
                # Combine the two lists into a list of tuples
                combined = list(zip(acc_ids, aslaas_nos))

                # Sort the combined list
                combined.sort()

                # Now, 'combined' is sorted based on the first element of each tuple
                # You can get the sorted 'acc_ids' and its corresponding 'aslaas_nos' like this:
                acc_ids, aslaas_nos = zip(*combined)

            self.account_ids = acc_ids
            self.aslaas_numbers = aslaas_nos

            # Retrieve the corresponding account numbers from the dictionary
            account_numbers = self.parent.dda.get_acc_nos_using_ids(acc_ids)

            self.account_nos = account_numbers

            # Combine the two lists into a list of tuples
            combined = list(zip(self.account_nos, self.aslaas_numbers))

            for data in combined:
                self.account_numbers_text_edit.append(f"{data[0]}:{data[1]}")
        except:
            self.parent.show_error_message("Error", "Error While Getting Update Details")

    def confirm_updates(self):
        try:
            acc_nos = self.account_nos
            aslaas_nos = self.aslaas_numbers
            acc_ids = self.account_ids

            self.update_aslaas_thread = AslaasUpdateThread(self.parent, acc_nos, aslaas_nos, acc_ids)
            self.update_aslaas_thread.finished_signal.connect(self.aslaas_update_completed)
            self.update_aslaas_thread.finished_with_error.connect(self.aslaas_update_error)
            self.update_aslaas_thread.start()
        except Exception as e :
            self.parent.show_error_message("Error", "Error while Updating Aslaas")

    def aslaas_update_completed(self):
        self.parent.show_popup_message("Message", "Aslaas Update and Sync Sucessful")

    def aslaas_update_error(self):
        self.parent.show_error_message("Error", "Error while Updating Aslaas")


# Download Report
class DownloadReportPage(QWidget):
    def __init__(self, parent):
        super().__init__()
        self.parent = parent
        self.init_ui()

    def init_ui(self):
        # Add widgets and layout for InterestCalculatorPage
        layout = QVBoxLayout()
        # Lot Reference Number Input
        self.lot_reference_label = QLabel("Lot Reference Number:")
        self.lot_reference_edit = QLineEdit()

        # Download Button
        self.download_button = QPushButton("Download Report")
        self.download_button.clicked.connect(self.download_report)

        layout.addWidget(self.lot_reference_label)
        layout.addWidget(self.lot_reference_edit)
        layout.addWidget(self.download_button)
        layout.addStretch()

        self.setLayout(layout)
        self.parent.stacked_widget.addWidget(self)

        self.parent.create_sidebar_button("Download Report", 4, "./_internal/static/download.svg")

    def download_report(self):
        try:
            referance_number = self.lot_reference_edit.text()
            if not referance_number=="":
                # Create and start the DownloadReportThread
                self.download_report_thread = DownloadReportThread(self.parent, referance_number)
                self.download_report_thread.finished_signal.connect(self.download_report_completed)
                self.download_report_thread.finished_with_error.connect(self.download_report_error)
                self.download_report_thread.start()
            else:
                self.parent.show_error_message('Error', 'Please enter a valid lot reference number')
        except Exception as e :
            self.download_report_error("Some Error Occoured !")

    def download_report_completed(self, result):
        # Add any additional UI updates or operations needed after the download completes
        self.parent.show_popup_message("Message", f"Downloaded report at: {result}")

    def download_report_error(self, error_message):
        self.parent.show_error_message("Error" , "Error While Downloading Report")


# Setup Basic Details
class SettingsPage(QWidget):
    def __init__(self, parent):
        super().__init__()
        self.parent = parent
        self.user_id_edit_mode = False
        self.pass_edit_mode = False
        self.download_folder_edit_mode = False
        self.agent_id_edit_mode = False
        self.agent_address_edit_mode = False
        self.lisence_exp_date_edit_mode = False
        self.agent_name_edit_mode = False
        self.agent_husband_name_edit_mode = False
        self.ocr_apikey_edit_mode = False
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()

        # Agent Name
        self.agent_name_label = QLabel("Agent Name:")
        self.agent_name_edit = QLineEdit(self.parent.agent_name)
        self.agent_name_edit.setReadOnly(True)

        self.agent_name_edit_button = QPushButton("Edit")
        self.agent_name_edit_button.clicked.connect(self.edit_agent_name)

        # Agent Husband Name
        self.agent_husband_name_label = QLabel("Agent Husband Name:")
        self.agent_husband_name_edit = QLineEdit(self.parent.agent_husband_name)
        self.agent_husband_name_edit.setReadOnly(True)

        self.agent_husband_name_edit_button = QPushButton("Edit")
        self.agent_husband_name_edit_button.clicked.connect(self.edit_agent_husband_name)


        # User ID
        self.user_id_label = QLabel("User ID:")
        self.user_id_edit = QLineEdit(self.parent.user_id)
        self.user_id_edit.setReadOnly(True)

        self.user_id_edit_button = QPushButton("Edit")
        self.user_id_edit_button.clicked.connect(self.edit_user_id)

        # Password
        self.password_label = QLabel("Password:")
        self.password_edit = QLineEdit(self.parent.user_password)
        self.password_edit.setEchoMode(QLineEdit.Password)
        self.password_edit.setReadOnly(True)

        self.password_edit_button = QPushButton("Edit")
        self.password_edit_button.clicked.connect(self.edit_password)

        self.show_password_checkbox = QCheckBox("Show Password")
        self.show_password_checkbox.stateChanged.connect(self.toggle_password_visibility)

        # Download Folder
        self.download_folder_label = QLabel("Default Download Folder:")
        self.download_folder_edit = QLineEdit(self.parent.def_download_dir)
        self.download_folder_edit.setReadOnly(True)

        self.download_folder_edit_button = QPushButton("Edit")
        self.download_folder_edit_button.clicked.connect(self.edit_download_folder)

        # Agent ID
        self.agent_id_label = QLabel("Agent ID:")
        self.agent_id_edit = QLineEdit(self.parent.agent_id)
        self.agent_id_edit.setReadOnly(True)

        self.agent_id_edit_button = QPushButton("Edit")
        self.agent_id_edit_button.clicked.connect(self.edit_agent_id)

        # Agent Address
        self.agent_address_label = QLabel("Agent Address:")
        self.agent_address_edit = QLineEdit(self.parent.agent_address)
        self.agent_address_edit.setReadOnly(True)

        self.agent_address_edit_button = QPushButton("Edit")
        self.agent_address_edit_button.clicked.connect(self.edit_agent_address)

        # License Expiration Date
        self.lisence_exp_date_label = QLabel("License Expiration Date:")
        self.lisence_exp_date_edit = QLineEdit(self.parent.lisence_exp_date)
        self.lisence_exp_date_edit.setReadOnly(True)

        self.lisence_exp_date_edit_button = QPushButton("Edit")
        self.lisence_exp_date_edit_button.clicked.connect(self.edit_lisence_exp_date)

        # OCR API Key
        self.ocr_apikey_label = QLabel("OCR API Key:")
        self.ocr_apikey_edit = QLineEdit(self.parent.ocr_apikey)
        self.ocr_apikey_edit.setReadOnly(True)

        self.ocr_apikey_edit_button = QPushButton("Edit")
        self.ocr_apikey_edit_button.clicked.connect(self.edit_ocr_apikey)

        # Sync Accounts
        self.sync_accounts_button = QPushButton("  Sync Accounts")
        self.sync_accounts_button.setIcon(QIcon("./_internal/static/sync.svg"))
        self.sync_accounts_button.clicked.connect(self.sync_accounts)

        # Sync Aslaas
        self.sync_aslaas_button = QPushButton("  Sync Aslaas")
        self.sync_aslaas_button.setIcon(QIcon("./_internal/static/sync.svg"))
        self.sync_aslaas_button.clicked.connect(self.sync_aslaas)

        # Agent Name
        agent_name_hbox = QHBoxLayout()
        agent_name_hbox.addWidget(self.agent_name_edit)
        agent_name_hbox.addWidget(self.agent_name_edit_button)
        agent_name_hbox.setAlignment(Qt.AlignLeft)
        agent_name_hbox.setContentsMargins(0, 0, 0, 0)
        agent_name_hbox.setSpacing(5)
        agent_name_hbox.setSizeConstraint(QLayout.SetMinimumSize)

        self.agent_name_edit_button.setFixedWidth(100)

        layout.addWidget(self.agent_name_label)
        layout.addLayout(agent_name_hbox)

        # Agent Husband Name
        agent_husband_name_hbox = QHBoxLayout()
        agent_husband_name_hbox.addWidget(self.agent_husband_name_edit)
        agent_husband_name_hbox.addWidget(self.agent_husband_name_edit_button)
        agent_husband_name_hbox.setAlignment(Qt.AlignLeft)
        agent_husband_name_hbox.setContentsMargins(0, 0, 0, 0)
        agent_husband_name_hbox.setSpacing(5)
        agent_husband_name_hbox.setSizeConstraint(QLayout.SetMinimumSize)

        self.agent_husband_name_edit_button.setFixedWidth(100)

        layout.addWidget(self.agent_husband_name_label)
        layout.addLayout(agent_husband_name_hbox)

        # User ID
        user_id_hbox = QHBoxLayout()
        user_id_hbox.addWidget(self.user_id_edit)
        user_id_hbox.addWidget(self.user_id_edit_button)
        user_id_hbox.setAlignment(Qt.AlignLeft)
        user_id_hbox.setContentsMargins(0, 0, 0, 0)
        user_id_hbox.setSpacing(5)
        user_id_hbox.setSizeConstraint(QLayout.SetMinimumSize)

        self.user_id_edit_button.setFixedWidth(100)

        layout.addWidget(self.user_id_label)
        layout.addLayout(user_id_hbox)

        # Password
        password_hbox = QHBoxLayout()
        password_hbox.addWidget(self.password_edit)
        password_hbox.addWidget(self.show_password_checkbox)
        password_hbox.addWidget(self.password_edit_button)
        password_hbox.setAlignment(Qt.AlignLeft)
        password_hbox.setContentsMargins(0, 0, 0, 0)
        password_hbox.setSpacing(10)
        password_hbox.setSizeConstraint(QLayout.SetMinimumSize)

        self.password_edit_button.setFixedWidth(100)

        layout.addWidget(self.password_label)
        layout.addLayout(password_hbox)

        # Download Folder
        download_folder_hbox = QHBoxLayout()
        download_folder_hbox.addWidget(self.download_folder_edit)
        download_folder_hbox.addWidget(self.download_folder_edit_button)
        download_folder_hbox.setAlignment(Qt.AlignLeft)
        download_folder_hbox.setContentsMargins(0, 0, 0, 0)
        download_folder_hbox.setSpacing(5)
        download_folder_hbox.setSizeConstraint(QLayout.SetMinimumSize)

        self.download_folder_edit_button.setFixedWidth(100)

        layout.addWidget(self.download_folder_label)
        layout.addLayout(download_folder_hbox)

        # Agent ID
        agent_id_hbox = QHBoxLayout()
        agent_id_hbox.addWidget(self.agent_id_edit)
        agent_id_hbox.addWidget(self.agent_id_edit_button)
        agent_id_hbox.setAlignment(Qt.AlignLeft)
        agent_id_hbox.setContentsMargins(0, 0, 0, 0)
        agent_id_hbox.setSpacing(5)
        agent_id_hbox.setSizeConstraint(QLayout.SetMinimumSize)

        self.agent_id_edit_button.setFixedWidth(100)

        layout.addWidget(self.agent_id_label)
        layout.addLayout(agent_id_hbox)

        # Agent Address
        agent_address_hbox = QHBoxLayout()
        agent_address_hbox.addWidget(self.agent_address_edit)
        agent_address_hbox.addWidget(self.agent_address_edit_button)
        agent_address_hbox.setAlignment(Qt.AlignLeft)
        agent_address_hbox.setContentsMargins(0, 0, 0, 0)
        agent_address_hbox.setSpacing(5)
        agent_address_hbox.setSizeConstraint(QLayout.SetMinimumSize)

        self.agent_address_edit_button.setFixedWidth(100)

        layout.addWidget(self.agent_address_label)
        layout.addLayout(agent_address_hbox)

        # License Expiration Date
        lisence_exp_date_hbox = QHBoxLayout()
        lisence_exp_date_hbox.addWidget(self.lisence_exp_date_edit)
        lisence_exp_date_hbox.addWidget(self.lisence_exp_date_edit_button)
        lisence_exp_date_hbox.setAlignment(Qt.AlignLeft)
        lisence_exp_date_hbox.setContentsMargins(0, 0, 0, 0)
        lisence_exp_date_hbox.setSpacing(5)
        lisence_exp_date_hbox.setSizeConstraint(QLayout.SetMinimumSize)

        self.lisence_exp_date_edit_button.setFixedWidth(100)

        layout.addWidget(self.lisence_exp_date_label)
        layout.addLayout(lisence_exp_date_hbox)

        # API Key
        ocr_apikey_hbox = QHBoxLayout()
        ocr_apikey_hbox.addWidget(self.ocr_apikey_edit)
        ocr_apikey_hbox.addWidget(self.ocr_apikey_edit_button)
        ocr_apikey_hbox.setAlignment(Qt.AlignLeft)
        ocr_apikey_hbox.setContentsMargins(0, 0, 0, 0)
        ocr_apikey_hbox.setSpacing(5)
        ocr_apikey_hbox.setSizeConstraint(QLayout.SetMinimumSize)

        self.ocr_apikey_edit_button.setFixedWidth(100)

        layout.addWidget(self.ocr_apikey_label)
        layout.addLayout(ocr_apikey_hbox)

        # Sync Accounts
        layout.addWidget(self.sync_accounts_button)

        # Sync Aslaas
        layout.addWidget(self.sync_aslaas_button)

        self.setLayout(layout)
        self.parent.stacked_widget.addWidget(self)

        #self.parent.load_settings()
        self.parent.create_sidebar_button("Settings", 5, "./_internal/static/settings.svg")



    def edit_agent_name(self):
        if self.agent_name_edit_mode:
            self.parent.agent_name = self.agent_name_edit.text()
            self.agent_name_edit.setReadOnly(True)
        else:
            self.agent_name_edit.setReadOnly(False)

        self.parent.save_settings()
        self.parent.initialize_assistants()
        self.toggle_agent_name_edit_mode()

    def edit_agent_husband_name(self):
        if self.agent_husband_name_edit_mode:
            self.parent.agent_husband_name = self.agent_husband_name_edit.text()
            self.agent_husband_name_edit.setReadOnly(True)
        else:
            self.agent_husband_name_edit.setReadOnly(False)

        self.parent.save_settings()
        self.parent.initialize_assistants()
        self.toggle_agent_husband_name_edit_mode()

    def edit_user_id(self):
        if self.user_id_edit_mode:
            self.parent.user_id = self.user_id_edit.text()
            self.user_id_edit.setReadOnly(True)
        else:
            self.user_id_edit.setReadOnly(False)

        self.parent.save_settings()
        self.parent.initialize_assistants()
        self.toggle_user_id_edit_mode()

    def edit_password(self):
        if self.pass_edit_mode:
            self.parent.user_password = self.password_edit.text()
            self.password_edit.setReadOnly(True)
        else:
            self.password_edit.setReadOnly(False)

        self.parent.save_settings()
        self.parent.initialize_assistants()
        self.toggle_pass_edit_mode()

    def edit_download_folder(self):
        folder_path = QFileDialog.getExistingDirectory(self, "Select Download Folder", self.parent.def_download_dir)
        if folder_path:
            self.parent.def_download_dir = folder_path
            self.download_folder_edit.setText(folder_path)

        self.parent.save_settings()
        self.parent.initialize_assistants()

    def edit_agent_id(self):
        if self.agent_id_edit_mode:
            self.parent.agent_id = self.agent_id_edit.text()
            self.agent_id_edit.setReadOnly(True)
        else:
            self.agent_id_edit.setReadOnly(False)

        self.parent.save_settings()
        self.parent.initialize_assistants()
        self.toggle_agent_id_edit_mode()

    def edit_agent_address(self):
        if self.agent_address_edit_mode:
            self.parent.agent_address = self.agent_address_edit.text()
            self.agent_address_edit.setReadOnly(True)
        else:
            self.agent_address_edit.setReadOnly(False)

        self.parent.save_settings()
        self.parent.initialize_assistants()
        self.toggle_agent_address_edit_mode()

    def edit_lisence_exp_date(self):
        if self.lisence_exp_date_edit_mode:
            self.parent.lisence_exp_date = self.lisence_exp_date_edit.text()
            self.lisence_exp_date_edit.setReadOnly(True)
        else:
            self.lisence_exp_date_edit.setReadOnly(False)

        self.parent.save_settings()
        self.parent.initialize_assistants()
        self.toggle_lisence_exp_date_edit_mode()

    def edit_ocr_apikey(self):
        if self.ocr_apikey_edit_mode:
            self.parent.ocr_apikey = self.ocr_apikey_edit.text()
            self.ocr_apikey_edit.setReadOnly(True)
        else:
            self.ocr_apikey_edit.setReadOnly(False)

        self.parent.save_settings()
        self.parent.initialize_assistants()
        self.toggle_ocr_apikey_edit_mode()

    def toggle_agent_name_edit_mode(self):
        self.agent_name_edit_mode = not self.agent_name_edit_mode
        self.agent_name_edit_button.setText("Edit" if not self.agent_name_edit_mode else "Save")

    def toggle_agent_husband_name_edit_mode(self):
        self.agent_husband_name_edit_mode = not self.agent_husband_name_edit_mode
        self.agent_husband_name_edit_button.setText("Edit" if not self.agent_husband_name_edit_mode else "Save")

    def toggle_user_id_edit_mode(self):
        self.user_id_edit_mode = not self.user_id_edit_mode
        self.user_id_edit_button.setText("Edit" if not self.user_id_edit_mode else "Save")

    def toggle_pass_edit_mode(self):
        self.pass_edit_mode = not self.pass_edit_mode
        self.password_edit_button.setText("Edit" if not self.pass_edit_mode else "Save")

    def toggle_download_folder_edit_mode(self):
        self.download_folder_edit_mode = not self.download_folder_edit_mode
        self.download_folder_edit_button.setText("Edit" if not self.download_folder_edit_mode else "Save")

    def toggle_agent_id_edit_mode(self):
        self.agent_id_edit_mode = not self.agent_id_edit_mode
        self.agent_id_edit_button.setText("Edit" if not self.agent_id_edit_mode else "Save")

    def toggle_agent_address_edit_mode(self):
        self.agent_address_edit_mode = not self.agent_address_edit_mode
        self.agent_address_edit_button.setText("Edit" if not self.agent_address_edit_mode else "Save")

    def toggle_lisence_exp_date_edit_mode(self):
        self.lisence_exp_date_edit_mode = not self.lisence_exp_date_edit_mode
        self.lisence_exp_date_edit_button.setText("Edit" if not self.lisence_exp_date_edit_mode else "Save")

    def toggle_ocr_apikey_edit_mode(self):
        self.ocr_apikey_edit_mode = not self.ocr_apikey_edit_mode
        self.ocr_apikey_edit_button.setText("Edit" if not self.ocr_apikey_edit_mode else "Save")

    def toggle_password_visibility(self, state):
        if state == 2:
            self.password_edit.setEchoMode(QLineEdit.Normal)
        else:
            self.password_edit.setEchoMode(QLineEdit.Password)

    def sync_accounts(self):
        try:
            self.sync_accounts_thread = SyncAccountsThread(self.parent)
            self.sync_accounts_thread.finished_signal.connect(self.sync_accounts_completed)
            self.sync_accounts_thread.finished_with_error.connect(self.sync_accounts_error)
            self.sync_accounts_thread.start()
        except Exception as e:
            self.parent.show_error_message("Error", "Error while syncing Accounts")

    def sync_aslaas(self):
        try:
            acc_nos = self.parent.dda.get_ac_nos_without_aslaas()
            aslaas_nos = ['APPLIEDFOR'] * len(acc_nos)
            acc_ids = self.parent.dda.get_acc_ids_using_nos(acc_nos)
            self.update_aslaas_thread = AslaasUpdateThread(self.parent, acc_nos, aslaas_nos, acc_ids)
            self.update_aslaas_thread.finished_signal.connect(self.aslaas_update_completed)
            self.update_aslaas_thread.finished_with_error.connect(self.aslaas_update_error)
            self.update_aslaas_thread.start()
        except Exception as e:
            self.parent.show_error_message("Error", "Error while syncing Aslaas")

    def aslaas_update_completed(self):
        self.parent.show_popup_message("Message", "Aslaas Update and Sync Sucessful")

    def aslaas_update_error(self):
        self.parent.show_error_message("Error", "Error while syncing Aslaas")

    def sync_accounts_completed(self):
        self.parent.show_popup_message("Message", "Accounts Sync Sucessful")

    def sync_accounts_error(self):
        self.parent.show_error_message("Error", "Error while syncing Accounts")


# About Developer
class AboutPage(QWidget):
    def __init__(self, parent):
        super().__init__()
        self.parent = parent
        self.init_ui()

    def init_ui(self):
        # Add widgets and layout for AboutPage
        layout = QVBoxLayout()
        layout.addStretch()
        layout.addWidget(QLabel("v2024.07.01"), alignment = Qt.AlignCenter)
        layout.addStretch()
        self.setLayout(layout)
        self.parent.stacked_widget.addWidget(self)

        self.parent.create_sidebar_button("About", 6, "./_internal/static/about.svg")


# Appearance Page
class AppearancePage(QWidget):
    def __init__(self, parent):
        super().__init__()
        self.parent = parent
        self.init_ui()
        self.update_theme()

    def init_ui(self):
        layout = QVBoxLayout()

        # Theme selection
        theme_label = QLabel("Theme:")
        self.theme_combo = QComboBox()
        self.theme_combo.addItems(["Light", "Dark"])
        self.theme_combo.setCurrentText(self.parent.theme.capitalize())
        self.theme_combo.currentIndexChanged.connect(self.update_theme)

        theme_layout = QHBoxLayout()
        theme_layout.addWidget(theme_label)
        theme_layout.addWidget(self.theme_combo)

        layout.addLayout(theme_layout)

        # Color selection with dropdown
        color_dropdown_label = QLabel("Accent Color:")
        self.color_combo = QComboBox()
        self.color_options = ['amber', 'blue', 'cyan', 'lightgreen', 'pink', 'purple', 'red', 'teal', 'yellow']
        for color in self.color_options:
            self.color_combo.addItem(color)
        self.color_combo.setCurrentText(self.parent.ascent)
        self.color_combo.currentIndexChanged.connect(self.update_theme)

        color_dropdown_layout = QHBoxLayout()
        color_dropdown_layout.addWidget(color_dropdown_label)
        color_dropdown_layout.addWidget(self.color_combo)

        layout.addLayout(color_dropdown_layout)

        # Scale selection
        scale_label = QLabel("Scale:")
        self.scale_combo = QComboBox()
        self.scale_options = [str(i) for i in range(-4, 3)]
        for scale in self.scale_options:
            self.scale_combo.addItem(scale)
        self.scale_combo.setCurrentText(str(self.parent.scale))
        self.scale_combo.currentIndexChanged.connect(self.update_theme)

        scale_layout = QHBoxLayout()
        scale_layout.addWidget(scale_label)
        scale_layout.addWidget(self.scale_combo)

        layout.addLayout(scale_layout)

        layout.addStretch()
        self.setLayout(layout)

    def update_theme(self):
        self.parent.theme = self.theme_combo.currentText().lower()
        self.parent.ascent = self.color_combo.currentText()
        self.parent.scale = int(self.scale_combo.currentText())  # Convert to integer
        extra = {
            'density_scale': self.parent.scale,
        }
        apply_stylesheet(app, theme=f"{self.parent.theme}_{self.parent.ascent}.xml", extra=extra, css_file='./_internal/static/styles.qss')
        self.parent.save_settings()


# Appearance Window as popup
class AppearanceWindow(QDialog):
    def __init__(self, parent):
        super().__init__(parent)
        self.setWindowTitle("Appearance Settings")
        self.parent = parent
        self.appearance_page = AppearancePage(self.parent)
        self.setLayout(QVBoxLayout())
        self.layout().addWidget(self.appearance_page)
        self.resize(400, 300)

        def closeEvent(self, event):
            self.parent.appearance_window = None
            event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = RDApplication()

    window.showMaximized()
    sys.exit(app.exec())
